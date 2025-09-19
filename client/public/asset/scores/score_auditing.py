import openpyxl
dse_subject_map = {}
dse_wb = openpyxl.load_workbook('DSE_subjects.xlsx')
dse_ws = dse_wb.active
for row in dse_ws.iter_rows(min_row=2, values_only=True):
    dse_subject_map[row[0]] = row[3]
wb = openpyxl.load_workbook('score.xlsx')
ws = wb.active


def name(subject):
    return dse_subject_map.get(subject, subject)


for row in ws.iter_rows(min_row=2, values_only=True):
    code = row[0]
    subject_compulsory = row[5]
    subject_optional_1 = row[6]
    subject_optional_2 = row[7]
    subject_free_number = row[8]
    subject_free_weight = row[9]
    subject_weight_limit = row[10]
    result = ""
    if subject_free_number is None:
        exit()
    print("\n", code)
    if subject_compulsory:
        subjects = eval(subject_compulsory)
        for subject in subjects:
            result += name(subject) + "*" + str(subjects[subject]) + " + "
    if subject_optional_1:
        subjects = eval(subject_optional_1)
        for subject in subjects:
            result += name(subject) + "*" + str(subjects[subject]) + "/"
        result += "\b + "
    if subject_optional_2:
        subjects = eval(subject_optional_2)
        for subject in subjects:
            result += name(subject) + "*" + str(subjects[subject]) + "/"
        result += "\b + "
    result += "最佳" + str(subject_free_number) + "科"
    if subject_weight_limit:
        result += f"\n其中{str(subject_weight_limit)}科可乘："
    else:
        result += "\n其中任意科可乘："
    if subject_free_weight:
        subjects = eval(subject_free_weight)
        for subject in subjects:
            result += name(subject) + "*" + str(subjects[subject]) + "/"
        result += "\b"
    print(result)
